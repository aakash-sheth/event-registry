from rest_framework import viewsets, status
from rest_framework import serializers as drf_serializers
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.conf import settings
from django.utils import timezone
import csv
from urllib.parse import quote
from .models import Event, RSVP, Guest, InvitePage, SubEvent, GuestSubEventInvite, WhatsAppTemplate
from .serializers import (
    EventSerializer, EventCreateSerializer,
    RSVPSerializer, RSVPCreateSerializer,
    GuestSerializer, GuestCreateSerializer,
    InvitePageSerializer, InvitePageCreateSerializer, InvitePageUpdateSerializer,
    SubEventSerializer, SubEventCreateSerializer,
    GuestSubEventInviteSerializer,
    WhatsAppTemplateSerializer
)
import re
from .utils import get_country_code, format_phone_with_country_code, normalize_csv_header, upload_to_s3, parse_phone_number
from apps.items.models import RegistryItem
from apps.items.serializers import RegistryItemSerializer


class EventViewSet(viewsets.ModelViewSet):
    serializer_class = EventSerializer
    
    def retrieve(self, request, *args, **kwargs):
        """Override retrieve"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_queryset(self):
        """Hosts can only see their own events - strict privacy enforcement"""
        # Try to include expiry_date if it exists, otherwise exclude it
        try:
            # First try with all fields including expiry_date
            return Event.objects.filter(host=self.request.user)
        except Exception as e:
            # If that fails (likely missing expiry_date column), try without it
            try:
                return Event.objects.filter(host=self.request.user).only(
                    'id', 'host_id', 'slug', 'title', 'event_type', 'date', 'city', 'country',
                    'is_public', 'has_rsvp', 'has_registry', 'banner_image', 'description',
                    'additional_photos', 'page_config', 'whatsapp_message_template',
                    'created_at', 'updated_at'
                )
            except Exception:
                # Last resort: return empty queryset to prevent 500 errors
                return Event.objects.none()
    
    def get_object(self):
        """Override to ensure host can only access their own events"""
        obj = super().get_object()
        # Double-check ownership (even though queryset is filtered)
        if obj.host != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access your own events.")
        return obj
    
    def get_serializer_class(self):
        if self.action == 'create':
            return EventCreateSerializer
        return EventSerializer
    
    def perform_create(self, serializer):
        """Ensure event is created with current user as host"""
        serializer.save(host=self.request.user)
    
    def create(self, request, *args, **kwargs):
        """Override create to add better error handling"""
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            # Log the full error for debugging
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating event: {str(e)}", exc_info=True)
            
            # Return a more helpful error message
            error_message = str(e)
            if 'has_registry' in error_message or 'has_rsvp' in error_message:
                return Response(
                    {'error': 'Database migration required. Please run: python manage.py migrate'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            elif 'slug' in error_message.lower():
                return Response(
                    {'error': f'Slug validation error: {error_message}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                return Response(
                    {'error': f'Failed to create event: {error_message}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
    
    def list(self, request, *args, **kwargs):
        """Override list to handle missing database fields gracefully"""
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            # If there's a database error (likely missing expiry_date column), 
            # try to serialize without it
            from rest_framework.response import Response
            from rest_framework import status
            
            # Try to get queryset without problematic fields
            try:
                queryset = self.get_queryset()
                # Manually serialize to avoid database errors
                serializer = self.get_serializer(queryset, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
            except Exception:
                # If that also fails, return empty list
                return Response([], status=status.HTTP_200_OK)
    
    def _verify_event_ownership(self, event):
        """Helper method to verify event ownership - raises PermissionDenied if not owner"""
        if not self.request.user.is_authenticated:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Authentication required.")
        if event.host != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access your own events and their data.")
        return True
    
    @action(detail=True, methods=['get'])
    def orders(self, request, id=None):
        """Get orders for this event - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        from apps.orders.models import Order
        from apps.orders.serializers import OrderSerializer
        
        # Only get orders for this specific event (already verified ownership)
        orders = Order.objects.filter(event=event).order_by('-created_at')
        serializer = OrderSerializer(orders, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def rsvps(self, request, id=None):
        """Get RSVPs for this event - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        # Only get active RSVPs for this specific event (exclude removed)
        rsvps = RSVP.objects.filter(event=event, is_removed=False).order_by('-created_at')
        serializer = RSVPSerializer(rsvps, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['delete'], url_path='rsvps/(?P<rsvp_id>[^/.]+)')
    def delete_rsvp(self, request, id=None, rsvp_id=None):
        """Remove an RSVP (soft delete) - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        try:
            rsvp = RSVP.objects.get(id=rsvp_id, event=event)
        except RSVP.DoesNotExist:
            return Response({'error': 'RSVP not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Soft delete RSVP (preserve historical data)
        rsvp.is_removed = True
        rsvp.save()
        return Response({
            'message': 'RSVP removed. The record is preserved but will not appear in active lists.',
            'soft_delete': True
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get', 'post'])
    def guests(self, request, id=None):
        """Get or create guests for this event - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        if request.method == 'GET':
            # Get all invited guests (we'll separate active/removed and sort in Python)
            all_guests = Guest.objects.filter(event=event).prefetch_related('rsvps')
            
            # Get all RSVPs for this event to help with sorting (exclude removed RSVPs)
            all_event_rsvps = RSVP.objects.filter(event=event, is_removed=False).select_related('guest')
            
            # Create a mapping of guest_id -> RSVP for sorting
            guest_rsvp_map = {}
            for rsvp in all_event_rsvps:
                if rsvp.guest_id:
                    guest_rsvp_map[rsvp.guest_id] = rsvp
            
            # Separate active and removed guests
            active_guests = []
            removed_guests = []
            
            for guest in all_guests:
                # Get RSVP for this guest (if exists) for sorting
                rsvp = guest_rsvp_map.get(guest.id)
                
                # Determine sort key: For guests with RSVP use RSVP.created_at, otherwise guest.created_at
                sort_date = rsvp.created_at if rsvp else guest.created_at
                
                guest_data = {
                    'guest': guest,
                    'rsvp': rsvp,
                    'sort_date': sort_date
                }
                
                if guest.is_removed:
                    removed_guests.append(guest_data)
                else:
                    active_guests.append(guest_data)
            
            # Sort: Primary by is_removed (False first), Secondary by sort_date (recent first)
            # Active guests: sort by sort_date descending (recent first)
            active_guests.sort(key=lambda x: x['sort_date'], reverse=True)
            # Removed guests: sort by sort_date descending (recent first)
            removed_guests.sort(key=lambda x: x['sort_date'], reverse=True)
            
            # Extract guest objects in sorted order
            sorted_active_guests = [g['guest'] for g in active_guests]
            sorted_removed_guests = [g['guest'] for g in removed_guests]
            
            guests_serializer = GuestSerializer(sorted_active_guests, many=True)
            removed_guests_serializer = GuestSerializer(sorted_removed_guests, many=True)
            
            # Get RSVPs that don't have a corresponding guest (other guests)
            # These are people who RSVP'd but weren't in the original guest list
            rsvps_without_guests = RSVP.objects.filter(
                event=event,
                guest__isnull=True,
                is_removed=False
            )
            
            # Also check for RSVPs where guest exists but phone doesn't match (edge case)
            # This handles cases where phone format differences prevent matching
            guest_phones = set(all_guests.values_list('phone', flat=True))
            other_rsvps = []
            removed_other_rsvps = []
            
            for rsvp in rsvps_without_guests:
                # Check if phone matches any guest (with various formats)
                rsvp_phone_digits = re.sub(r'\D', '', rsvp.phone)
                matches_guest = False
                for guest_phone in guest_phones:
                    guest_phone_digits = re.sub(r'\D', '', guest_phone)
                    if rsvp_phone_digits == guest_phone_digits:
                        matches_guest = True
                        break
                
                if not matches_guest:
                    if rsvp.is_removed:
                        removed_other_rsvps.append(rsvp)
                    else:
                        other_rsvps.append(rsvp)
            
            # Sort other RSVPs: Primary by is_removed (False first), Secondary by created_at (recent first)
            # Active other RSVPs: sort by created_at descending
            other_rsvps.sort(key=lambda x: x.created_at, reverse=True)
            # Removed other RSVPs: sort by created_at descending
            removed_other_rsvps.sort(key=lambda x: x.created_at, reverse=True)
            
            other_guests_serializer = RSVPSerializer(other_rsvps, many=True)
            removed_other_guests_serializer = RSVPSerializer(removed_other_rsvps, many=True)
            
            return Response({
                'guests': guests_serializer.data,
                'removed_guests_list': removed_guests_serializer.data,  # Removed invited guests
                'other_guests': other_guests_serializer.data,
                'removed_guests': removed_other_guests_serializer.data  # Removed other guests (RSVPs)
            })
        
        elif request.method == 'POST':
            # Bulk create from list
            guests_data = request.data.get('guests', [])
            if not isinstance(guests_data, list):
                return Response({'error': 'guests must be a list'}, status=status.HTTP_400_BAD_REQUEST)
            
            created_guests = []
            errors = []
            event_country_code = get_country_code(event.country)
            
            for idx, guest_data in enumerate(guests_data):
                serializer = GuestCreateSerializer(data=guest_data)
                if serializer.is_valid():
                    phone = serializer.validated_data.get('phone', '')
                    # Format phone with country code if not already formatted
                    if phone and not phone.startswith('+'):
                        country_code = guest_data.get('country_code') or event_country_code
                        phone = format_phone_with_country_code(phone, country_code)
                    
                    # Check if guest with this phone already exists
                    existing_guest = Guest.objects.filter(event=event, phone=phone).first()
                    if existing_guest:
                        errors.append(f"Guest {idx + 1}: Phone {phone} already exists (Name: {existing_guest.name})")
                        continue
                    # Event ownership already verified, safe to create
                    try:
                        guest_data['phone'] = phone
                        # Remove country_code from data before creating (keep country_iso)
                        country_iso = guest_data.pop('country_iso', None)
                        guest_data.pop('country_code', None)
                        guest = Guest.objects.create(
                            event=event,
                            country_iso=country_iso or '',
                            **{k: v for k, v in guest_data.items() if k not in ['country_code', 'country_iso']}
                        )
                        created_guests.append(guest)
                    except Exception as e:
                        errors.append(f"Guest {idx + 1}: {str(e)}")
                else:
                    errors.append(f"Guest {idx + 1}: {serializer.errors}")
            
            response_data = GuestSerializer(created_guests, many=True).data
            if errors:
                response_data = {'created': response_data, 'errors': errors}
            
            status_code = status.HTTP_201_CREATED if created_guests else status.HTTP_400_BAD_REQUEST
            return Response(response_data, status=status_code)
    
    @action(detail=True, methods=['post'], url_path='guests/import')
    def import_guests_csv(self, request, id=None):
        """Import guests from CSV or Excel file - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        import csv
        import io
        import os
        
        file = request.FILES['file']
        file_name = file.name.lower()
        file_extension = os.path.splitext(file_name)[1].lower()
        
        # Determine file type and read accordingly
        rows = []
        if file_extension in ['.xlsx', '.xls']:
            # Excel file
            try:
                from openpyxl import load_workbook
                
                # Read Excel file
                file.seek(0)  # Reset file pointer
                workbook = load_workbook(file, read_only=True, data_only=True)
                sheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    headers.append(str(cell.value).strip() if cell.value else '')
                
                # Read data rows
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header:  # Only add non-empty headers
                            value = row[idx] if idx < len(row) else None
                            row_dict[header] = str(value).strip() if value is not None else ''
                    if any(row_dict.values()):  # Only add non-empty rows
                        rows.append((row_num, row_dict))
                
                workbook.close()
            except ImportError:
                return Response({'error': 'Excel support requires openpyxl library. Please install it.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            except Exception as e:
                return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        elif file_extension == '.csv':
            # CSV file
            try:
                file.seek(0)  # Reset file pointer
                decoded_file = file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
                    rows.append((row_num, row))
            except Exception as e:
                return Response({'error': f'Failed to read CSV file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': f'Unsupported file format: {file_extension}. Please use CSV (.csv) or Excel (.xlsx, .xls) files.'}, status=status.HTTP_400_BAD_REQUEST)
        
        created_count = 0
        errors = []
        
        event_country_code = get_country_code(event.country)
        
        # Define required columns (case-insensitive matching)
        required_columns = {'name', 'phone', 'email', 'relationship', 'notes', 'is_removed'}
        
        # Track custom fields metadata (normalized key -> display label)
        custom_fields_metadata = event.custom_fields_metadata.copy() if event.custom_fields_metadata else {}
        reserved_variables = {'name', 'event_title', 'event_date', 'event_url', 'host_name', 'event_location', 'map_direction'}
        
        # Process first row to identify custom columns
        if rows:
            first_row = rows[0][1]  # Get the row dict
            for original_header in first_row.keys():
                # Normalize header
                normalized_header = original_header.strip().lower()
                # Skip if it's a required column (case-insensitive)
                if normalized_header in required_columns:
                    continue
                
                # Normalize to variable format
                normalized_key = normalize_csv_header(original_header)
                
                # Skip empty normalized keys
                if not normalized_key:
                    continue
                
                # Handle collisions with reserved variables
                final_key = normalized_key
                if final_key in reserved_variables:
                    # Add suffix to avoid collision
                    counter = 1
                    while f"{final_key}_custom" in custom_fields_metadata or f"{final_key}_custom" in reserved_variables:
                        final_key = f"{normalized_key}_{counter}"
                        counter += 1
                    if final_key == normalized_key:
                        final_key = f"{normalized_key}_custom"
                
                # Store metadata
                if final_key not in custom_fields_metadata:
                    custom_fields_metadata[final_key] = {
                        'display_label': original_header.strip(),
                        'source': 'csv_import',
                    }
        
        # Update event's custom_fields_metadata
        if custom_fields_metadata:
            event.custom_fields_metadata = custom_fields_metadata
            event.save(update_fields=['custom_fields_metadata'])
        
        for row_num, row in rows:
            try:
                name = row.get('name') or row.get('Name') or row.get('NAME')
                phone = row.get('phone') or row.get('Phone') or row.get('PHONE')
                
                if not name or not name.strip():
                    errors.append(f"Row {row_num}: Name is required")
                    continue
                
                if not phone or not phone.strip():
                    errors.append(f"Row {row_num}: Phone number is required")
                    continue
                
                # Clean phone number
                phone_clean = phone.strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                phone_digits = ''.join(filter(str.isdigit, phone_clean))
                
                # If phone is exactly 10 digits, add event's country code
                if len(phone_digits) == 10:
                    phone = f"{event_country_code}{phone_digits}"
                else:
                    # Format phone with country code (handles various formats)
                    phone = format_phone_with_country_code(phone.strip(), event_country_code)
                
                # Validate phone format (should start with + and have digits)
                if not phone.startswith('+') or len(phone.replace('+', '')) < 10:
                    errors.append(f"Row {row_num}: Invalid phone number format: {phone}")
                    continue
                
                # Parse is_removed column (if present)
                is_removed = False
                is_removed_str = (row.get('is_removed') or row.get('Is Removed') or row.get('IS_REMOVED') or row.get('is removed') or '').strip().lower()
                if is_removed_str in ['true', '1', 'yes', 'y', 'removed']:
                    is_removed = True
                
                # Check if guest with this phone already exists for this event (exclude removed guests from duplicate check)
                existing_guest = Guest.objects.filter(event=event, phone=phone, is_removed=False).first()
                if existing_guest:
                    # Guest with this phone already exists - skip and report
                    errors.append(f"Row {row_num}: Guest with phone {phone} already exists (Name: {existing_guest.name})")
                    continue
                
                # Extract custom fields
                custom_fields = {}
                for original_header, value in row.items():
                    normalized_header = original_header.strip().lower()
                    # Skip required columns
                    if normalized_header in required_columns:
                        continue
                    
                    # Normalize to variable format
                    normalized_key = normalize_csv_header(original_header)
                    if not normalized_key:
                        continue
                    
                    # Find the final key (handling collisions)
                    final_key = normalized_key
                    if final_key in reserved_variables:
                        # Check if we have a custom version
                        if f"{normalized_key}_custom" in custom_fields_metadata:
                            final_key = f"{normalized_key}_custom"
                        else:
                            # Find the numbered version
                            counter = 1
                            while f"{normalized_key}_{counter}" not in custom_fields_metadata and f"{normalized_key}_{counter}" in reserved_variables:
                                counter += 1
                            if f"{normalized_key}_{counter}" in custom_fields_metadata:
                                final_key = f"{normalized_key}_{counter}"
                            else:
                                final_key = f"{normalized_key}_custom"
                    
                    # Only store if this is a known custom field
                    if final_key in custom_fields_metadata:
                        # Store the value (strip whitespace)
                        value_str = str(value).strip() if value else ''
                        if value_str:
                            custom_fields[final_key] = value_str
                
                # Generate guest token
                guest = Guest.objects.create(
                    event=event,
                    name=name.strip(),
                    phone=phone,
                    country_iso='',  # CSV import doesn't have country ISO, can be updated later
                    email=(row.get('email') or row.get('Email') or row.get('EMAIL') or '').strip() or None,
                    relationship=(row.get('relationship') or row.get('Relationship') or '').strip() or '',
                    notes=(row.get('notes') or row.get('Notes') or '').strip() or '',
                    is_removed=is_removed,
                    custom_fields=custom_fields,
                )
                
                # Generate guest token if not present
                if not guest.guest_token:
                    guest.generate_guest_token()
                
                created_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return Response({
            'created': created_count,
            'errors': errors if errors else None
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['put', 'patch'], url_path='design')
    def update_design(self, request, id=None):
        """Update event page design - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        # Handle page_config (Living Poster template)
        if 'page_config' in request.data:
            page_config = request.data.get('page_config')
            if isinstance(page_config, dict):
                event.page_config = page_config
                # Only update page_config and updated_at fields (faster than full save)
                event.save(update_fields=['page_config', 'updated_at'])
                
                # Sync to InvitePage.config if it exists (ensures invite page has latest config)
                try:
                    invite_page = InvitePage.objects.filter(event=event).first()
                    if invite_page:
                        invite_page.config = page_config
                        invite_page.save(update_fields=['config', 'updated_at'])
                except Exception:
                    # InvitePage might not exist yet - ignore
                    pass
                
                # Return minimal response instead of full event object (much faster)
                return Response({
                    'status': 'success',
                    'message': 'Design saved successfully'
                }, status=status.HTTP_200_OK)
        
        # Legacy form-based fields (for backward compatibility)
        banner_image = request.data.get('banner_image', '')
        description = request.data.get('description', '')
        additional_photos = request.data.get('additional_photos', [])
        
        # Validate additional_photos (max 5)
        if isinstance(additional_photos, list) and len(additional_photos) > 5:
            return Response(
                {'error': 'Maximum 5 additional photos allowed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Update event - track which fields need updating
        update_fields = []
        if banner_image is not None:
            event.banner_image = banner_image
            update_fields.append('banner_image')
        if description is not None:
            event.description = description
            update_fields.append('description')
        if additional_photos is not None:
            event.additional_photos = additional_photos if isinstance(additional_photos, list) else []
            update_fields.append('additional_photos')
        
        # Only save if there are fields to update
        if update_fields:
            update_fields.append('updated_at')
            event.save(update_fields=update_fields)
        
        # Return minimal response instead of full event object (much faster)
        return Response({
            'status': 'success',
            'message': 'Design saved successfully'
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['put', 'patch'], url_path='guests/(?P<guest_id>[^/.]+)')
    def update_guest(self, request, id=None, guest_id=None):
        """Update a guest in the list - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        try:
            guest = Guest.objects.get(id=guest_id, event=event)
        except Guest.DoesNotExist:
            return Response({'error': 'Guest not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = GuestCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Update guest fields
        guest.name = serializer.validated_data.get('name', guest.name)
        guest.email = serializer.validated_data.get('email', guest.email)
        guest.relationship = serializer.validated_data.get('relationship', guest.relationship)
        guest.notes = serializer.validated_data.get('notes', guest.notes)
        
        # Handle phone update (format with country code)
        if 'phone' in serializer.validated_data:
            country_code = serializer.validated_data.get('country_code') or get_country_code(event.country)
            phone = serializer.validated_data['phone']
            guest.phone = format_phone_with_country_code(phone, country_code)
        
        if 'country_iso' in serializer.validated_data:
            guest.country_iso = serializer.validated_data['country_iso']
        
        guest.save()
        return Response(GuestSerializer(guest).data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['delete'], url_path='guests/(?P<guest_id>[^/.]+)')
    def delete_guest(self, request, id=None, guest_id=None):
        """Delete or remove a guest from the list - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        try:
            guest = Guest.objects.get(id=guest_id, event=event)
        except Guest.DoesNotExist:
            return Response({'error': 'Guest not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Check if guest has any valid RSVP (will_attend in 'yes', 'no', 'maybe')
        rsvp = RSVP.objects.filter(
            event=event,
            guest=guest,
            will_attend__in=['yes', 'no', 'maybe'],
            is_removed=False
        ).first()
        
        # Also check by phone number (in case guest wasn't linked during RSVP creation)
        if not rsvp:
            rsvp = RSVP.objects.filter(
                event=event,
                phone=guest.phone,
                will_attend__in=['yes', 'no', 'maybe'],
                is_removed=False
            ).first()
        
        if rsvp:
            # Guest has a valid RSVP - soft delete only (cannot hard delete)
            guest.is_removed = True
            guest.save()
            return Response({
                'message': 'Guest removed (soft delete). They have an RSVP and cannot be permanently deleted.',
                'guest_id': guest.id,
                'is_removed': True,
                'soft_delete': True
            }, status=status.HTTP_200_OK)
        else:
            # Guest has no valid RSVP - hard delete allowed
            guest.delete()
            return Response({
                'message': 'Guest deleted successfully',
                'soft_delete': False
            }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], url_path='guests/(?P<guest_id>[^/.]+)/reinstate')
    def reinstate_guest(self, request, id=None, guest_id=None):
        """Reinstate a removed guest - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        try:
            guest = Guest.objects.get(id=guest_id, event=event)
        except Guest.DoesNotExist:
            return Response({'error': 'Guest not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Reinstate guest
        guest.is_removed = False
        guest.save()
        return Response({
            'message': 'Guest reinstated successfully.',
            'guest': GuestSerializer(guest).data
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], url_path='rsvps/(?P<rsvp_id>[^/.]+)/reinstate')
    def reinstate_rsvp(self, request, id=None, rsvp_id=None):
        """Reinstate a removed RSVP - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        try:
            rsvp = RSVP.objects.get(id=rsvp_id, event=event)
        except RSVP.DoesNotExist:
            return Response({'error': 'RSVP not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Reinstate RSVP
        rsvp.is_removed = False
        rsvp.save()
        return Response({
            'message': 'RSVP reinstated successfully.',
            'rsvp': RSVPSerializer(rsvp).data
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], url_path='share/whatsapp')
    def share_whatsapp(self, request, id=None):
        """Generate WhatsApp share links for event - host only"""
        event = self.get_object()
        self._verify_event_ownership(event)
        
        share_type = request.data.get('type', 'public')  # 'public', 'guest', 'bulk'
        guest_id = request.data.get('guest_id')
        guest_ids = request.data.get('guest_ids', [])
        custom_message = request.data.get('message', '')
        
        frontend_origin = settings.FRONTEND_ORIGIN
        event_url = f"{frontend_origin}/invite/{event.slug}"
        
        # Default message template
        if not custom_message:
            date_str = event.date.strftime('%B %d, %Y') if event.date else 'TBD'
            custom_message = f"Hey! 💛\n\nJust wanted to share {event.title} on {date_str}!\n\nPlease confirm here: {event_url}\n\n- {event.host.name or 'Your Host'}"
        
        if share_type == 'public':
            # Generate public share link
            message = quote(custom_message)
            whatsapp_url = f"https://wa.me/?text={message}"
            return Response({
                'whatsapp_url': whatsapp_url,
                'message': custom_message,
                'event_url': event_url,
            }, status=status.HTTP_200_OK)
        
        elif share_type == 'guest':
            # Generate link for specific guest
            if not guest_id:
                return Response({'error': 'guest_id is required for guest sharing'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                guest = Guest.objects.get(id=guest_id, event=event)
            except Guest.DoesNotExist:
                return Response({'error': 'Guest not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Format phone number (remove + and any spaces)
            phone = guest.phone.replace('+', '').replace(' ', '').replace('-', '')
            
            # Personalize message with guest name
            personalized_message = custom_message.replace('Hey!', f"Hey {guest.name}!")
            message = quote(personalized_message)
            whatsapp_url = f"https://wa.me/{phone}/?text={message}"
            
            return Response({
                'whatsapp_url': whatsapp_url,
                'message': personalized_message,
                'event_url': event_url,
                'guest': {
                    'id': guest.id,
                    'name': guest.name,
                    'phone': guest.phone,
                }
            }, status=status.HTTP_200_OK)
        
        elif share_type == 'bulk':
            # Bulk share - generate links for multiple guests
            guest_ids = request.data.get('guest_ids', [])
            if not guest_ids:
                return Response({'error': 'guest_ids is required for bulk sharing'}, status=status.HTTP_400_BAD_REQUEST)
            
            results = []
            for guest_id in guest_ids:
                try:
                    guest = Guest.objects.get(id=guest_id, event=event)
                    if not guest.phone:
                        results.append({
                            'guest_id': guest_id,
                            'error': 'Guest has no phone number'
                        })
                        continue
                    
                    # Format phone number
                    phone = guest.phone.replace('+', '').replace(' ', '').replace('-', '')
                    
                    # Personalize message
                    personalized_message = custom_message.replace('Hey!', f"Hey {guest.name}!")
                    message = quote(personalized_message)
                    whatsapp_url = f"https://wa.me/{phone}/?text={message}"
                    
                    results.append({
                        'guest_id': guest.id,
                        'guest_name': guest.name,
                        'whatsapp_url': whatsapp_url,
                        'message': personalized_message,
                    })
                except Guest.DoesNotExist:
                    results.append({
                        'guest_id': guest_id,
                        'error': 'Guest not found'
                    })
            
            return Response({
                'results': results,
                'event_url': event_url,
            }, status=status.HTTP_200_OK)
        
        else:
            return Response({'error': 'Invalid share_type'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'], url_path='whatsapp-preview')
    def whatsapp_preview(self, request, id=None):
        """Preview WhatsApp message with template and guest data"""
        event = self.get_object()
        self._verify_event_ownership(event)
        
        template_id = request.data.get('template_id')
        guest_id = request.data.get('guest_id')
        raw_body = request.data.get('raw_body')
        
        if not template_id and not raw_body:
            return Response({'error': 'Either template_id or raw_body is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get base URL
        base_url = request.build_absolute_uri('/').rstrip('/')
        
        # Get template if provided
        template_text = None
        if template_id:
            try:
                template = WhatsAppTemplate.objects.get(id=template_id, event=event)
                template_text = template.template_text
            except WhatsAppTemplate.DoesNotExist:
                return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            template_text = raw_body
        
        # Get guest if provided
        guest = None
        if guest_id:
            try:
                guest = Guest.objects.get(id=guest_id, event=event)
            except Guest.DoesNotExist:
                return Response({'error': 'Guest not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Render template
        from .utils import render_template_with_guest
        rendered_message, warnings = render_template_with_guest(
            template_text,
            event,
            guest,
            base_url
        )
        
        return Response({
            'preview': rendered_message,
            'warnings': warnings,
        })
    
    @action(detail=True, methods=['get'], url_path='whatsapp-templates/available-variables')
    def get_available_variables(self, request, id=None):
        """Get list of available variables for this event (default + custom from CSV)"""
        event = self.get_object()
        self._verify_event_ownership(event)
        
        variables = []
        
        # Default variables
        default_vars = [
            {'key': '[name]', 'label': 'Guest Name', 'description': 'Name of the guest', 'example': 'Sarah', 'is_custom': False},
            {'key': '[event_title]', 'label': 'Event Title', 'description': 'Title of the event', 'example': event.title, 'is_custom': False},
            {'key': '[event_date]', 'label': 'Event Date', 'description': 'Date of the event', 'example': event.date.strftime('%B %d, %Y') if event.date else 'TBD', 'is_custom': False},
            {'key': '[event_url]', 'label': 'Event URL', 'description': 'Link to the event invitation', 'example': f'https://example.com/invite/{event.slug}', 'is_custom': False},
            {'key': '[host_name]', 'label': 'Host Name', 'description': 'Name of the event host', 'example': event.host.name or 'Host', 'is_custom': False},
            {'key': '[event_location]', 'label': 'Event Location', 'description': 'Location of the event', 'example': event.city or 'Location TBD', 'is_custom': False},
            {'key': '[map_direction]', 'label': 'Map Direction Link', 'description': 'Google Maps link to event location', 'example': 'https://maps.google.com/?q=Location', 'is_custom': False},
        ]
        variables.extend(default_vars)
        
        # Custom variables from CSV imports
        custom_metadata = event.custom_fields_metadata or {}
        for normalized_key, metadata in custom_metadata.items():
            if isinstance(metadata, dict):
                display_label = metadata.get('display_label', normalized_key)
                example = metadata.get('example', '—')
            else:
                # Backward compatibility
                display_label = metadata
                example = '—'
            
            variables.append({
                'key': f'[{normalized_key}]',
                'label': display_label,
                'description': f'Custom field from CSV: {display_label}',
                'example': example,
                'is_custom': True,
            })
        
        return Response({'variables': variables})
    
    @action(detail=True, methods=['get'], url_path='system-default-template')
    def get_system_default_template(self, request, id=None):
        """Get the system default template"""
        # Verify event ownership (required by permission_classes)
        event = self.get_object()
        self._verify_event_ownership(event)
        
        # System default template is global, not event-specific
        try:
            system_template = WhatsAppTemplate.objects.filter(is_system_default=True).first()
            if not system_template:
                return Response({'error': 'System default template not found'}, status=status.HTTP_404_NOT_FOUND)
            
            return Response(WhatsAppTemplateSerializer(system_template).data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            for guest in guests:
                phone = guest.phone.replace('+', '').replace(' ', '').replace('-', '')
                personalized_message = custom_message.replace('Hey!', f"Hey {guest.name}!")
                message = quote(personalized_message)
                whatsapp_url = f"https://wa.me/{phone}/?text={message}"
                
                links.append({
                    'guest_id': guest.id,
                    'guest_name': guest.name,
                    'guest_phone': guest.phone,
                    'whatsapp_url': whatsapp_url,
                    'message': personalized_message,
                })
            
            return Response({
                'links': links,
                'event_url': event_url,
                'count': len(links),
            }, status=status.HTTP_200_OK)
        
        else:
            return Response({'error': 'Invalid share type. Use "public", "guest", or "bulk"'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'], url_path='impact')
    def get_impact(self, request, id=None):
        """Get impact metrics for expired event - host only"""
        event = self.get_object()
        self._verify_event_ownership(event)
        
        from .utils import calculate_event_impact
        impact = calculate_event_impact(event)
        
        if impact is None:
            return Response({
                'error': 'Event is not expired yet. Impact can only be calculated for expired events.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(impact, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'], url_path='impact/overall')
    def get_overall_impact(self, request):
        """Get overall impact across all expired events - host only"""
        from .utils import calculate_event_impact
        from datetime import date
        
        # Get all events for this host
        # Use only() to avoid selecting expiry_date if it doesn't exist
        try:
            events = Event.objects.filter(host=request.user).only(
                'id', 'host_id', 'slug', 'title', 'event_type', 'date', 'city', 'country',
                'is_public', 'has_rsvp', 'has_registry', 'banner_image', 'description',
                'additional_photos', 'page_config',
                'created_at', 'updated_at'
            )
        except Exception:
            # If only() fails, try without it
            try:
                events = Event.objects.filter(host=request.user)
            except Exception:
                # If query fails completely, return empty result
                return Response({
                    'total_plates_saved': 0,
                    'total_paper_saved': 0,
                    'total_gifts_received': 0,
                    'total_gift_value_rupees': 0,
                    'total_paper_saved_on_gifts': 0,
                    'expired_events_count': 0,
                    'events': []
                }, status=status.HTTP_200_OK)
        
        # Safely check if expired (handles case where expiry_date field might not exist)
        expired_events = []
        for e in events:
            try:
                if e.is_expired:
                    expired_events.append(e)
            except Exception:
                # If there's an error accessing is_expired, skip this event
                continue
        
        if not expired_events:
            return Response({
                'total_plates_saved': 0,
                'total_paper_saved': 0,
                'total_gifts_received': 0,
                'total_gift_value_rupees': 0,
                'total_paper_saved_on_gifts': 0,
                'expired_events_count': 0,
                'events': []
            }, status=status.HTTP_200_OK)
        
        # Aggregate impact across all expired events
        total_plates_saved = 0
        total_paper_saved = 0
        total_gifts_received = 0
        total_gift_value_paise = 0
        total_paper_saved_on_gifts = 0
        
        events_impact = []
        
        for event in expired_events:
            impact = calculate_event_impact(event)
            if impact:
                total_plates_saved += impact['food_saved']['plates_saved']
                total_paper_saved += impact['paper_saved']['web_rsvps']
                total_gifts_received += impact['gifts_received']['total_gifts']
                total_gift_value_paise += impact['gifts_received']['total_value_paise']
                total_paper_saved_on_gifts += impact['paper_saved_on_gifts']['cash_gifts']
                
                events_impact.append({
                    'event_id': event.id,
                    'event_title': event.title,
                    'event_date': event.date.isoformat() if event.date else None,
                    'expiry_date': event.expiry_date.isoformat() if event.expiry_date else None,
                    'impact': impact
                })
        
        return Response({
            'total_plates_saved': total_plates_saved,
            'total_paper_saved': total_paper_saved,
            'total_gifts_received': total_gifts_received,
            'total_gift_value_rupees': total_gift_value_paise / 100,
            'total_paper_saved_on_gifts': total_paper_saved_on_gifts,
            'expired_events_count': len(expired_events),
            'events': events_impact
        }, status=status.HTTP_200_OK)


class InvitePageViewSet(viewsets.ModelViewSet):
    """
    Invite page management - host only, privacy protected
    """
    serializer_class = InvitePageSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_queryset(self):
        """Hosts can only see invite pages for their own events"""
        return InvitePage.objects.filter(event__host=self.request.user)
    
    def get_object(self):
        """Override to retrieve invite page by event_id instead of id"""
        # If event_id is in kwargs (from URL), get invite page by event
        event_id = self.kwargs.get('event_id')
        if event_id:
            event = get_object_or_404(Event, id=event_id, host=self.request.user)
            try:
                invite_page = InvitePage.objects.get(event=event)
            except InvitePage.DoesNotExist:
                # Raise Http404 so DRF returns proper 404 response
                from django.http import Http404
                raise Http404("Invite page not found for this event")
            # Verify ownership
            if invite_page.event.host != self.request.user:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("You can only access invite pages for your own events.")
            return invite_page
        
        # Otherwise, use default behavior (lookup by id)
        obj = super().get_object()
        if obj.event.host != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access invite pages for your own events.")
        return obj
    
    def get_serializer_class(self):
        if self.action == 'create':
            return InvitePageCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return InvitePageUpdateSerializer
        return InvitePageSerializer
    
    def create(self, request, *args, **kwargs):
        """Create invite page for an event"""
        event_id = self.kwargs.get('event_id')
        if not event_id:
            return Response(
                {'error': 'event_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        event = get_object_or_404(Event, id=event_id, host=request.user)
        
        # Check if invite page already exists
        if InvitePage.objects.filter(event=event).exists():
            return Response(
                {'error': 'Invite page already exists for this event'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        invite_page = serializer.save(event=event)
        return Response(InvitePageSerializer(invite_page).data, status=status.HTTP_201_CREATED)
    
    def retrieve(self, request, *args, **kwargs):
        """Retrieve invite page by event_id"""
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except InvitePage.DoesNotExist:
            return Response(
                {'error': 'Invite page not found for this event'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            # If get_object fails (e.g., invite page doesn't exist), return 404
            if 'DoesNotExist' in str(type(e).__name__):
                return Response(
                    {'error': 'Invite page not found for this event'},
                    status=status.HTTP_404_NOT_FOUND
                )
            raise
    
    @action(detail=True, methods=['post'], url_path='publish')
    def publish(self, request, id=None):
        """Publish/unpublish invite page"""
        invite_page = self.get_object()
        is_published = request.data.get('is_published', True)
        invite_page.is_published = is_published
        invite_page.save()
        return Response(InvitePageSerializer(invite_page).data, status=status.HTTP_200_OK)


class PublicInviteViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Public invite page view - no authentication required
    Supports guest token via ?g=<token> query parameter for guest-scoped rendering
    """
    serializer_class = InvitePageSerializer
    permission_classes = [AllowAny]
    lookup_field = 'slug'
    
    def get_queryset(self):
        """Only return published invite pages"""
        return InvitePage.objects.filter(is_published=True)
    
    def retrieve(self, request, *args, **kwargs):
        """Retrieve invite page with guest-scoped sub-events if token provided"""
        slug = kwargs.get('slug')
        
        # DEBUG: Comprehensive logging for production investigation
        import logging
        import time
        import sys
        logger = logging.getLogger(__name__)
        start_time = time.time()
        
        # Log to both logger and stdout for visibility
        # Only log to stdout in production if LOG_LEVEL is DEBUG to reduce overhead
        log_to_stdout = os.environ.get('LOG_LEVEL', 'INFO') == 'DEBUG'
        def debug_log(msg, level='INFO'):
            """Log to both logger and stdout for debugging"""
            log_msg = f"[PublicInviteViewSet] {msg}"
            if level == 'INFO':
                logger.info(log_msg)
            elif level == 'WARNING':
                logger.warning(log_msg)
            elif level == 'ERROR':
                logger.error(log_msg)
            elif level == 'DEBUG':
                logger.debug(log_msg)
            # Only print to stdout if explicitly enabled (reduces overhead in production)
            if log_to_stdout or level in ['WARNING', 'ERROR']:
                print(log_msg, file=sys.stdout, flush=True)
        
        debug_log(f"🔍 START: Retrieving invite page for slug: {slug}")
        debug_log(f"Request method: {request.method}, Path: {request.path}")
        debug_log(f"Query params: {dict(request.query_params)}")
        
        # Try to get invite page by slug with optimized query (case-insensitive)
        query_start = time.time()
        debug_log(f"Step 1: Looking up InvitePage with slug='{slug}' and is_published=True")
        try:
            # Try exact match first (most common case)
            invite_page = InvitePage.objects.select_related('event').get(slug=slug, is_published=True)
            event = invite_page.event
            query_time = time.time() - query_start
            debug_log(f"✅ Step 1 SUCCESS: Found invite page (query took {query_time:.3f}s)")
            debug_log(f"   InvitePage ID: {invite_page.id}, Slug: {invite_page.slug}, Published: {invite_page.is_published}")
            debug_log(f"   Event ID: {event.id}, Event Slug: {event.slug}, Title: {event.title}")
        except InvitePage.DoesNotExist:
            query_time = time.time() - query_start
            debug_log(f"❌ Step 1 FAILED: No exact match found (query took {query_time:.3f}s)")
            debug_log(f"Step 2: Trying case-insensitive match with slug__iexact='{slug}'")
            query_start = time.time()
            try:
                invite_page = InvitePage.objects.select_related('event').get(slug__iexact=slug, is_published=True)
                event = invite_page.event
                query_time = time.time() - query_start
                debug_log(f"✅ Step 2 SUCCESS: Found with case-insensitive match (query took {query_time:.3f}s)")
                debug_log(f"   InvitePage Slug: {invite_page.slug} (requested: {slug})")
                debug_log(f"   Event ID: {event.id}, Event Slug: {event.slug}")
            except InvitePage.DoesNotExist:
                query_time = time.time() - query_start
                debug_log(f"❌ Step 2 FAILED: No case-insensitive match found (query took {query_time:.3f}s)")
                debug_log(f"Step 3: Trying exact match without is_published check")
                query_start = time.time()
                try:
                    invite_page = InvitePage.objects.select_related('event').get(slug=slug)
                    event = invite_page.event
                    query_time = time.time() - query_start
                    debug_log(f"✅ Step 3 SUCCESS: Found unpublished invite page (query took {query_time:.3f}s)")
                    debug_log(f"   InvitePage ID: {invite_page.id}, Published: {invite_page.is_published}")
                    if not invite_page.is_published:
                        debug_log(f"⚠️  Publishing invite page...")
                        publish_start = time.time()
                        invite_page.is_published = True
                        invite_page.save(update_fields=['is_published'])
                        publish_time = time.time() - publish_start
                        debug_log(f"✅ Published (took {publish_time:.3f}s)")
                    debug_log(f"   Event ID: {event.id}, Event Slug: {event.slug}")
                except InvitePage.DoesNotExist:
                    query_time = time.time() - query_start
                    debug_log(f"❌ Step 3 FAILED: No exact match without published check (query took {query_time:.3f}s)")
                    debug_log(f"Step 4: Trying case-insensitive match without is_published check")
                    query_start = time.time()
                    try:
                        invite_page = InvitePage.objects.select_related('event').get(slug__iexact=slug)
                        event = invite_page.event
                        query_time = time.time() - query_start
                        debug_log(f"✅ Step 4 SUCCESS: Found unpublished with case-insensitive (query took {query_time:.3f}s)")
                        debug_log(f"   InvitePage Slug: {invite_page.slug} (requested: {slug}), Published: {invite_page.is_published}")
                        if not invite_page.is_published:
                            debug_log(f"⚠️  Publishing invite page...")
                            publish_start = time.time()
                            invite_page.is_published = True
                            invite_page.save(update_fields=['is_published'])
                            publish_time = time.time() - publish_start
                            debug_log(f"✅ Published (took {publish_time:.3f}s)")
                        debug_log(f"   Event ID: {event.id}, Event Slug: {event.slug}")
                    except InvitePage.DoesNotExist:
                        query_time = time.time() - query_start
                        debug_log(f"❌ Step 4 FAILED: No case-insensitive match without published check (query took {query_time:.3f}s)")
                        debug_log(f"Step 5: InvitePage not found, trying Event lookup with slug='{slug}'")
                        query_start = time.time()
                        # If invite page doesn't exist, try to find event by slug (case-insensitive)
                        try:
                            # Try exact match first
                            event = Event.objects.only('id', 'slug', 'page_config', 'event_structure', 'title', 'description', 'date', 'has_rsvp', 'has_registry').get(slug=slug)
                            query_time = time.time() - query_start
                            debug_log(f"✅ Step 5 SUCCESS: Found event (query took {query_time:.3f}s)")
                            debug_log(f"   Event ID: {event.id}, Slug: {event.slug}, Title: {event.title}")
                        except Event.DoesNotExist:
                            query_time = time.time() - query_start
                            debug_log(f"❌ Step 5 FAILED: No exact event match (query took {query_time:.3f}s)")
                            debug_log(f"Step 6: Trying case-insensitive event match")
                            query_start = time.time()
            try:
                                event = Event.objects.only('id', 'slug', 'page_config', 'event_structure', 'title', 'description', 'date', 'has_rsvp', 'has_registry').get(slug__iexact=slug)
                                query_time = time.time() - query_start
                                debug_log(f"✅ Step 6 SUCCESS: Found event with case-insensitive match (query took {query_time:.3f}s)")
                                debug_log(f"   Event Slug: {event.slug} (requested: {slug})")
                            except Event.DoesNotExist:
                                query_time = time.time() - query_start
                                debug_log(f"❌ Step 6 FAILED: Event not found (query took {query_time:.3f}s)", 'ERROR')
                                debug_log(f"❌ FINAL ERROR: Event not found for slug: {slug} (exact or case-insensitive)", 'ERROR')
                                
                                # Log 404 to CloudWatch with full request details for alerting
                                full_url = request.build_absolute_uri()
                                logger.error(
                                    f"INVITE_404: Invite page or event not found",
                                    extra={
                                        'event_type': 'invite_404',
                                        'slug': slug,
                                        'full_url': full_url,
                                        'path': request.path,
                                        'query_params': dict(request.query_params),
                                        'user_agent': request.META.get('HTTP_USER_AGENT', ''),
                                        'referer': request.META.get('HTTP_REFERER', ''),
                                        'remote_addr': request.META.get('REMOTE_ADDR', ''),
                                        'request_method': request.method,
                                    }
                                )
                                
                                from rest_framework.exceptions import NotFound
                                raise NotFound(f"Invite page or event not found for slug: {slug}")
                        
                        # Create a default invite page if it doesn't exist (optimized query)
                        debug_log(f"Step 7: Creating/getting InvitePage for event_id={event.id}")
                        create_start = time.time()
                        try:
                invite_page, created = InvitePage.objects.get_or_create(
                    event=event,
                    defaults={
                        'slug': event.slug,  # Use event slug as invite page slug
                        'is_published': True,  # Auto-publish if created
                        'config': event.page_config if event.page_config else {}  # Use event's page_config if available
                    }
                )
                            create_time = time.time() - create_start
                            if created:
                                debug_log(f"✅ Step 7 SUCCESS: Created new invite page (took {create_time:.3f}s)")
                                debug_log(f"   InvitePage ID: {invite_page.id}, Slug: {invite_page.slug}, Published: {invite_page.is_published}")
                            else:
                                debug_log(f"✅ Step 7 SUCCESS: Found existing invite page (took {create_time:.3f}s)")
                                debug_log(f"   InvitePage ID: {invite_page.id}, Slug: {invite_page.slug}, Published: {invite_page.is_published}")
                            
                # If it already existed but wasn't published, publish it
                if not created and not invite_page.is_published:
                                debug_log(f"⚠️  Publishing existing invite page...")
                                publish_start = time.time()
                    invite_page.is_published = True
                    invite_page.save(update_fields=['is_published'])
                                publish_time = time.time() - publish_start
                                debug_log(f"✅ Published (took {publish_time:.3f}s)")
                            
                # Always sync config from event.page_config if it exists and is more complete
                # This ensures invite page always has the latest design settings
                            # NOTE: Removed complex config comparison logic to improve performance
                            # Config sync should be handled by admin/API, not during public page loads
                if event.page_config and isinstance(event.page_config, dict) and len(event.page_config) > 0:
                                if not invite_page.config or not isinstance(invite_page.config, dict) or len(invite_page.config) == 0:
                                    debug_log(f"⚠️  Syncing config from event to invite page...")
                                    sync_start = time.time()
                        invite_page.config = event.page_config
                        invite_page.save(update_fields=['config'])
                                    sync_time = time.time() - sync_start
                                    debug_log(f"✅ Config synced (took {sync_time:.3f}s)")
                        except Exception as e:
                            debug_log(f"❌ ERROR creating/updating invite page for event {event.slug}: {e}", 'ERROR')
                            logger.error(f"[PublicInviteViewSet] Error creating/updating invite page for event {event.slug}: {e}", exc_info=True)
                            # Re-raise to return proper error to client
                            raise
        
        # Extract guest token from query params
        debug_log(f"Step 8: Processing guest token and sub-events")
        sub_events_start = time.time()
        guest_token = request.query_params.get('g', '').strip()
        guest = None
        allowed_sub_events = []
        
        if guest_token:
            debug_log(f"   Guest token provided: {guest_token[:10]}...")
            # Resolve guest token with optimized query
            try:
                guest_query_start = time.time()
                guest = Guest.objects.only('id', 'name', 'event_id', 'guest_token').get(
                    guest_token=guest_token, 
                    event=event, 
                    is_removed=False
                )
                guest_query_time = time.time() - guest_query_start
                debug_log(f"✅ Guest found (query took {guest_query_time:.3f}s): ID={guest.id}, Name={guest.name}")
                # Get allowed sub-events via join table with optimized query
                sub_events_query_start = time.time()
                allowed_sub_events = SubEvent.objects.filter(
                    guest_invites__guest=guest,
                    is_removed=False
                ).only('id', 'title', 'start_at', 'end_at', 'location', 'description', 'image_url', 'rsvp_enabled').order_by('start_at')
                sub_events_query_time = time.time() - sub_events_query_start
                debug_log(f"✅ Sub-events query completed (took {sub_events_query_time:.3f}s)")
            except Guest.DoesNotExist:
                debug_log(f"❌ Guest not found for token", 'WARNING')
                # Invalid token - return empty sub-events (guest won't see any)
                allowed_sub_events = SubEvent.objects.none()
                guest = None
        else:
            debug_log(f"   No guest token - using public sub-events")
            # Public link - only show public-visible sub-events with optimized query
            sub_events_query_start = time.time()
            allowed_sub_events = SubEvent.objects.filter(
                event=event,
                is_public_visible=True,
                is_removed=False
            ).only('id', 'title', 'start_at', 'end_at', 'location', 'description', 'image_url', 'rsvp_enabled').order_by('start_at')
            sub_events_query_time = time.time() - sub_events_query_start
            debug_log(f"✅ Public sub-events query completed (took {sub_events_query_time:.3f}s)")
        
        # Convert to list early to evaluate queryset and check count efficiently
        list_start = time.time()
        sub_events_list = list(allowed_sub_events)
        list_time = time.time() - list_start
        has_sub_events = len(sub_events_list) > 0
        debug_log(f"✅ Converted sub-events to list (took {list_time:.3f}s): {len(sub_events_list)} sub-events")
            
            # If we found sub-events but event structure is still SIMPLE, upgrade it
        # Use update() for better performance (single query instead of get + save)
        # Only do this for public links (not guest tokens) to avoid unnecessary writes
        if has_sub_events and not guest_token and event.event_structure == 'SIMPLE':
            debug_log(f"⚠️  Upgrading event structure from SIMPLE to ENVELOPE...")
            upgrade_start = time.time()
            Event.objects.filter(id=event.id, event_structure='SIMPLE').update(
                event_structure='ENVELOPE',
                updated_at=timezone.now()
            )
            # Refresh event object for serializer
            event.refresh_from_db(fields=['event_structure', 'updated_at'])
            upgrade_time = time.time() - upgrade_start
            debug_log(f"✅ Event structure upgraded (took {upgrade_time:.3f}s)")
        
        # Serialize sub-events (use list to avoid re-evaluating queryset)
        debug_log(f"Step 9: Serializing sub-events and guest context")
        serialize_start = time.time()
        serialized_sub_events = SubEventSerializer(sub_events_list, many=True).data
        serialize_time = time.time() - serialize_start
        debug_log(f"✅ Sub-events serialized (took {serialize_time:.3f}s): {len(serialized_sub_events)} items")
        
        # Serialize guest context only if guest exists (avoid unnecessary serialization)
        guest_context = None
        if guest:
            guest_serialize_start = time.time()
            guest_context = GuestSerializer(guest).data
            guest_serialize_time = time.time() - guest_serialize_start
            debug_log(f"✅ Guest context serialized (took {guest_serialize_time:.3f}s)")
        
        # Serialize with context
        debug_log(f"Step 10: Serializing final response")
        final_serialize_start = time.time()
        serializer = self.get_serializer(invite_page, context={
            'allowed_sub_events': serialized_sub_events,
            'guest_context': guest_context
        })
        final_serialize_time = time.time() - final_serialize_start
        debug_log(f"✅ Final serialization completed (took {final_serialize_time:.3f}s)")
        
        elapsed_time = time.time() - start_time
        sub_events_total_time = time.time() - sub_events_start
        
        # Always log timing summary for slow requests (important for debugging 504s)
        if elapsed_time > 1.0:
            debug_log(f"⚠️  SLOW REQUEST: {elapsed_time:.2f}s (sub-events: {sub_events_total_time:.2f}s)", 'WARNING')
        elif elapsed_time > 0.5:
            debug_log(f"📊 Request completed: {elapsed_time:.2f}s (sub-events: {sub_events_total_time:.2f}s)")
        # Only log detailed summary if DEBUG mode or slow request
        if log_to_stdout or elapsed_time > 1.0:
            debug_log(f"📊 TIMING SUMMARY:")
            debug_log(f"   Total request time: {elapsed_time:.3f}s")
            debug_log(f"   Sub-events processing: {sub_events_total_time:.3f}s")
            debug_log(f"🏁 END: Returning response for slug: {slug}")
        
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'], url_path='guests.csv')
    def guests_csv(self, request, id=None):
        """Export guest list (RSVPs) as CSV - host only, privacy protected"""
        event = self.get_object()
        self._verify_event_ownership(event)  # Explicit ownership check
        
        # Get all RSVPs for this event
        rsvps = RSVP.objects.filter(event=event).order_by('-created_at')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="guest_list_{event.slug}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Name', 
            'Phone', 
            'Email', 
            'Will Attend', 
            'Guests Count', 
            'Source Channel', 
            'Notes', 
            'RSVP Date',
            'Is Removed'
        ])
        
        for rsvp in rsvps:
            writer.writerow([
                rsvp.name,
                rsvp.phone,
                rsvp.email or '',
                rsvp.get_will_attend_display(),
                rsvp.guests_count,
                rsvp.get_source_channel_display(),
                rsvp.notes or '',
                rsvp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Yes' if rsvp.is_removed else 'No',
            ])
        
        return response


class PublicEventViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Public registry view - no authentication required
    
    PRIVACY NOTE: This viewset only exposes:
    - Public event details (title, date, city, slug)
    - Active registry items (no guest list, no RSVPs, no orders)
    - No host information beyond what's in EventSerializer
    
    All private data (guest lists, RSVPs, orders) is ONLY accessible to the event host
    through EventViewSet with authentication.
    """
    serializer_class = EventSerializer
    permission_classes = [AllowAny]
    lookup_field = 'slug'
    
    def get_queryset(self):
        """Return all events - privacy is enforced per-endpoint"""
        # Use select_related to avoid N+1 queries when accessing host.name
        return Event.objects.select_related('host').all()
    
    @action(detail=True, methods=['get'])
    def items(self, request, slug=None):
        """Get active items for public registry - no private data exposed"""
        event = get_object_or_404(Event, slug=slug)
        
        # For private events, verify user is in guest list
        if not event.is_public:
            phone = request.query_params.get('phone', '').strip()
            if not phone:
                return Response(
                    {'error': 'This is a private event. Phone number required to verify access.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Format phone with country code if not already formatted
            event_country_code = get_country_code(event.country)
            if not phone.startswith('+'):
                country_code = request.query_params.get('country_code', event_country_code)
                phone = format_phone_with_country_code(phone, country_code)
            
            # Try to find in guest list
            guest = None
            phone_digits_only = re.sub(r'\D', '', phone)
            provided_country_code = request.query_params.get('country_code', event_country_code)
            
            # First try exact phone match
            guest = Guest.objects.filter(event=event, phone=phone).first()
            
            # If not found, try matching by digits only
            if not guest:
                all_guests = Guest.objects.filter(event=event)
                for g in all_guests:
                    guest_phone_digits = re.sub(r'\D', '', g.phone)
                    if guest_phone_digits == phone_digits_only:
                        guest = g
                        break
                    
                    # Try matching last 10 digits with country code verification
                    if len(phone_digits_only) >= 10 and len(guest_phone_digits) >= 10:
                        local_number = phone_digits_only[-10:]
                        if guest_phone_digits.endswith(local_number):
                            stored_country_code, _ = parse_phone_number(g.phone)
                            if stored_country_code == provided_country_code:
                                guest = g
                                break
            
            if not guest:
                return Response(
                    {'error': 'This is a private event. Only invited guests can view the registry.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Check if registry is enabled for this event
        if not event.has_registry:
            return Response(
                {'error': 'Gift registry is not available for this event'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Only return active items - no guest list, no RSVPs, no order details
        items = RegistryItem.objects.filter(event=event, status='active').order_by('priority_rank', 'name')
        
        # Calculate remaining quantities
        items_data = []
        for item in items:
            remaining = item.qty_total - item.qty_purchased
            item_dict = RegistryItemSerializer(item).data
            item_dict['remaining'] = remaining
            items_data.append(item_dict)
        
        return Response({
            'event': EventSerializer(event).data,
            'items': items_data,
        })


@api_view(['GET'])
@permission_classes([AllowAny])
def get_rsvp(request, event_id):
    """Get existing RSVP by phone number (public endpoint)"""
    try:
        event = get_object_or_404(Event, id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Check if RSVP is enabled for this event
    if not event.has_rsvp:
        return Response(
            {'error': 'RSVP is not available for this event'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    phone = request.query_params.get('phone', '').strip()
    if not phone:
        return Response({'error': 'Phone number is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Format phone with country code if not already formatted
    event_country_code = get_country_code(event.country)
    original_phone = phone
    if phone and not phone.startswith('+'):
        country_code = request.query_params.get('country_code', event_country_code)
        phone = format_phone_with_country_code(phone, country_code)
    
    # Find existing RSVP - try multiple formats (grandfather clause: check RSVP first)
    existing_rsvp = RSVP.objects.filter(event=event, phone=phone, is_removed=False).first()
    
    # If not found with formatted phone, try variations
    if not existing_rsvp:
        # Try without + prefix
        phone_without_plus = phone.lstrip('+')
        existing_rsvp = RSVP.objects.filter(event=event, phone=phone_without_plus, is_removed=False).first()
        
        # Try with different country code formats (if provided country code doesn't match)
        if not existing_rsvp:
            provided_country_code = request.query_params.get('country_code', event_country_code)
            phone_digits_only = re.sub(r'\D', '', phone)
            
            # Try all RSVPs and check if the phone digits match (ignoring formatting, exclude removed)
            all_rsvps = RSVP.objects.filter(event=event, is_removed=False)
            for rsvp in all_rsvps:
                rsvp_phone_digits = re.sub(r'\D', '', rsvp.phone)
                
                # Check if phone digits match exactly
                if rsvp_phone_digits == phone_digits_only:
                    existing_rsvp = rsvp
                    break
                
                # If digits don't match exactly, try matching last 10 digits with country code verification
                if len(phone_digits_only) >= 10 and len(rsvp_phone_digits) >= 10:
                    local_number = phone_digits_only[-10:]
                    if rsvp_phone_digits.endswith(local_number):
                        # Extract country code from stored phone
                        stored_country_code, _ = parse_phone_number(rsvp.phone)
                        # Only match if country codes are the same
                        if stored_country_code == provided_country_code:
                            existing_rsvp = rsvp
                            break
    
    if existing_rsvp:
        # Check if linked guest is removed (block access)
        if existing_rsvp.guest and existing_rsvp.guest.is_removed:
            return Response(
                {'error': 'This guest has been removed. You cannot access your RSVP.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Return RSVP data (grandfather clause: existing RSVP allows access)
        rsvp_data = RSVPSerializer(existing_rsvp).data
        rsvp_data['found_in'] = 'rsvp'
        return Response(rsvp_data, status=status.HTTP_200_OK)
    else:
        # For both public and private events, check guest list if RSVP not found (exclude removed guests)
        phone_digits_only = re.sub(r'\D', '', phone)
        provided_country_code = request.query_params.get('country_code', event_country_code)
        
        # Try to find in guest list (exclude removed guests)
        guest = None
        # First try exact phone match
        guest = Guest.objects.filter(event=event, phone=phone, is_removed=False).first()
        
        # If not found, try matching by digits only
        if not guest:
            all_guests = Guest.objects.filter(event=event, is_removed=False)
            for g in all_guests:
                guest_phone_digits = re.sub(r'\D', '', g.phone)
                if guest_phone_digits == phone_digits_only:
                    guest = g
                    break
                
                # Try matching last 10 digits with country code verification
                if len(phone_digits_only) >= 10 and len(guest_phone_digits) >= 10:
                    local_number = phone_digits_only[-10:]
                    if guest_phone_digits.endswith(local_number):
                        stored_country_code, _ = parse_phone_number(g.phone)
                        if stored_country_code == provided_country_code:
                            guest = g
                            break
        
        if guest:
            # Found in guest list but no RSVP - return guest info
            guest_data = GuestSerializer(guest).data
            guest_data['found_in'] = 'guest_list'
            guest_data['has_rsvp'] = False
            return Response(guest_data, status=status.HTTP_200_OK)
        
        # Not found in RSVP or guest list
        debug_info = {}
        if settings.DEBUG:
            all_phones = list(RSVP.objects.filter(event=event).values_list('phone', flat=True))
            debug_info = {
                'searched_phone': phone,
                'original_phone': original_phone,
                'all_phones_in_db': all_phones,
                'total_rsvps': RSVP.objects.filter(event=event).count(),
            }
        return Response({
            'error': 'No RSVP or guest found for this phone number',
            'debug': debug_info
        }, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([AllowAny])
def check_phone_for_rsvp(request, event_id):
    """Stage 0: Check if phone number exists in guest list for private events"""
    try:
        event = get_object_or_404(Event, id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Check if RSVP is enabled for this event
    if not event.has_rsvp:
        return Response({'error': 'RSVP is not enabled for this event'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Only required for private events
    if event.is_public:
        return Response({'error': 'This endpoint is only for private events'}, status=status.HTTP_400_BAD_REQUEST)
    
    phone = request.data.get('phone', '').strip()
    country_code = request.data.get('country_code', '')
    
    if not phone:
        return Response({'error': 'Phone number is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Format phone with country code
    from .utils import format_phone_with_country_code, get_country_code, parse_phone_number
    event_country_code = get_country_code(event.country)
    original_phone = phone
    
    if phone and not phone.startswith('+'):
        if country_code:
            phone = format_phone_with_country_code(phone, country_code)
        else:
            phone = format_phone_with_country_code(phone, event_country_code)
    
    phone_digits_only = re.sub(r'\D', '', phone)
    provided_country_code = country_code or event_country_code
    
    # GRANDFATHER CLAUSE: First check if there's an existing RSVP (even if not in guest list)
    # This handles people who RSVP'd when event was public but later became private
    existing_rsvp = None
    # Try exact phone match first
    existing_rsvp = RSVP.objects.filter(event=event, phone=phone, is_removed=False).first()
    
    # If not found, try matching by digits only
    if not existing_rsvp:
        all_rsvps = RSVP.objects.filter(event=event, is_removed=False)
        for rsvp in all_rsvps:
            rsvp_phone_digits = re.sub(r'\D', '', rsvp.phone)
            if rsvp_phone_digits == phone_digits_only:
                existing_rsvp = rsvp
                break
            
            # Try matching last 10 digits with country code verification
            if len(phone_digits_only) >= 10 and len(rsvp_phone_digits) >= 10:
                local_number = phone_digits_only[-10:]
                if rsvp_phone_digits.endswith(local_number):
                    stored_country_code, _ = parse_phone_number(rsvp.phone)
                    if stored_country_code == provided_country_code:
                        existing_rsvp = rsvp
                        break
    
    # If existing RSVP found (grandfather clause), return RSVP data
    if existing_rsvp:
        # Check if linked guest is removed (block access)
        if existing_rsvp.guest and existing_rsvp.guest.is_removed:
            return Response(
                {'error': 'This guest has been removed. You cannot access your RSVP.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Return RSVP data for pre-filling form
        from .serializers import RSVPSerializer
        rsvp_data = RSVPSerializer(existing_rsvp).data
        rsvp_data['found_in'] = 'rsvp'
        rsvp_data['phone_verified'] = True
        return Response(rsvp_data, status=status.HTTP_200_OK)
    
    # If no existing RSVP, check guest list (exclude removed guests)
    guest = None
    # First try exact phone match
    guest = Guest.objects.filter(event=event, phone=phone, is_removed=False).first()
    
    # If not found, try matching by digits only
    if not guest:
        all_guests = Guest.objects.filter(event=event, is_removed=False)
        for g in all_guests:
            guest_phone_digits = re.sub(r'\D', '', g.phone)
            if guest_phone_digits == phone_digits_only:
                guest = g
                break
            
            # Try matching last 10 digits with country code verification
            if len(phone_digits_only) >= 10 and len(guest_phone_digits) >= 10:
                local_number = phone_digits_only[-10:]
                if guest_phone_digits.endswith(local_number):
                    stored_country_code, _ = parse_phone_number(g.phone)
                    if stored_country_code == provided_country_code:
                        guest = g
                        break
    
    if not guest:
        return Response(
            {'error': 'Phone number not found. Please try a different number or contact the host.'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Return guest data for pre-filling form
    from .serializers import GuestSerializer
    guest_data = GuestSerializer(guest).data
    guest_data['found_in'] = 'guest_list'
    guest_data['phone_verified'] = True
    
    return Response(guest_data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([AllowAny])
def create_rsvp(request, event_id):
    """Create or update RSVP for an event (public endpoint)"""
    try:
        event = get_object_or_404(Event, id=event_id)
    except Event.DoesNotExist:
        return Response({'error': 'Event not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Check if RSVP is enabled for this event
    if not event.has_rsvp:
        return Response(
            {'error': 'RSVP is not available for this event'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    serializer = RSVPCreateSerializer(data=request.data)
    if serializer.is_valid():
        phone = serializer.validated_data['phone']
        name = serializer.validated_data['name']
        
        # Format phone with country code if not already formatted
        event_country_code = get_country_code(event.country)
        if phone and not phone.startswith('+'):
            country_code = request.data.get('country_code') or event_country_code
            phone = format_phone_with_country_code(phone, country_code)
        
        # Check if RSVP already exists for this phone FIRST (grandfather clause)
        existing_rsvp = RSVP.objects.filter(event=event, phone=phone, is_removed=False).first()
        
        # For private events, check existing RSVP first (grandfather clause)
        if not event.is_public:
            if existing_rsvp:
                # Existing RSVP found - check if linked guest is removed
                if existing_rsvp.guest and existing_rsvp.guest.is_removed:
                    return Response(
                        {'error': 'This guest has been removed. You cannot update your RSVP.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
                # If not removed, proceed to update existing RSVP (grandfather clause allows access)
                guest = existing_rsvp.guest  # Use existing guest link
            else:
                # No existing RSVP, check guest list (exclude removed guests)
                phone_digits_only = re.sub(r'\D', '', phone)
                provided_country_code = request.data.get('country_code') or event_country_code
                
                # Try to find in guest list (exclude removed guests)
                guest = None
                # First try exact phone match
                guest = Guest.objects.filter(event=event, phone=phone, is_removed=False).first()
                
                # If not found, try matching by digits only
                if not guest:
                    all_guests = Guest.objects.filter(event=event, is_removed=False)
                for g in all_guests:
                    guest_phone_digits = re.sub(r'\D', '', g.phone)
                    if guest_phone_digits == phone_digits_only:
                        guest = g
                        break
                    
                    # Try matching last 10 digits with country code verification
                    if len(phone_digits_only) >= 10 and len(guest_phone_digits) >= 10:
                        local_number = phone_digits_only[-10:]
                        if guest_phone_digits.endswith(local_number):
                            stored_country_code, _ = parse_phone_number(g.phone)
                            if stored_country_code == provided_country_code:
                                guest = g
                                break
            
            if not guest:
                return Response(
                    {'error': 'This is a private event. Only invited guests can RSVP.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        else:
            # For public events, check existing RSVP first (grandfather clause)
            if existing_rsvp:
                # Existing RSVP found - check if linked guest is removed
                if existing_rsvp.guest and existing_rsvp.guest.is_removed:
                    return Response(
                        {'error': 'This guest has been removed. You cannot update your RSVP.'},
                        status=status.HTTP_403_FORBIDDEN
                    )
                # If not removed, proceed to update existing RSVP
                guest = existing_rsvp.guest  # Use existing guest link
            else:
                # For public events, try to match guest (optional linking, exclude removed)
                guest = None
                if event.guest_list.exists():
                    # Try to match by phone first, then by name (exclude removed)
                    guest = Guest.objects.filter(event=event, phone=phone, is_removed=False).first()
                    if not guest:
                        guest = Guest.objects.filter(event=event, name__iexact=name, is_removed=False).first()
        
        # Update phone in validated_data and remove country_code (not a model field)
        serializer.validated_data['phone'] = phone
        selected_sub_event_ids = serializer.validated_data.pop('selectedSubEventIds', [])
        rsvp_data = {k: v for k, v in serializer.validated_data.items() if k not in ['country_code', 'selectedSubEventIds']}
        
        # Handle ENVELOPE events with sub-events
        if event.event_structure == 'ENVELOPE':
            if event.rsvp_mode == 'PER_SUBEVENT':
                # PER_SUBEVENT mode: Create/update RSVP for each selected sub-event
                if not selected_sub_event_ids:
                    return Response(
                        {'error': 'selectedSubEventIds is required for PER_SUBEVENT mode'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Verify all sub-events belong to this event and are not removed
                sub_events = SubEvent.objects.filter(
                    id__in=selected_sub_event_ids,
                    event=event,
                    is_removed=False,
                    rsvp_enabled=True
                )
                
                if sub_events.count() != len(selected_sub_event_ids):
                    return Response(
                        {'error': 'Some sub-events not found, disabled, or do not belong to this event'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # For private events, verify guest has access to these sub-events
                if not event.is_public and guest:
                    allowed_sub_event_ids = set(
                        GuestSubEventInvite.objects.filter(guest=guest)
                        .values_list('sub_event_id', flat=True)
                    )
                    if not all(se_id in allowed_sub_event_ids for se_id in selected_sub_event_ids):
                        return Response(
                            {'error': 'You are not invited to some of the selected sub-events'},
                            status=status.HTTP_403_FORBIDDEN
                        )
                
                # Create/update RSVP for each sub-event
                created_rsvps = []
                for sub_event in sub_events:
                    # Check if RSVP already exists for this sub-event
                    existing_sub_rsvp = RSVP.objects.filter(
                        event=event,
                        phone=phone,
                        sub_event=sub_event,
                        is_removed=False
                    ).first()
                    
                    if existing_sub_rsvp:
                        # Update existing RSVP
                        for key, value in rsvp_data.items():
                            setattr(existing_sub_rsvp, key, value)
                        if guest:
                            existing_sub_rsvp.guest = guest
                        existing_sub_rsvp.save()
                        created_rsvps.append(existing_sub_rsvp)
                    else:
                        # Create new RSVP
                        rsvp = RSVP.objects.create(
                            event=event,
                            sub_event=sub_event,
                            guest=guest,
                            **rsvp_data
                        )
                        created_rsvps.append(rsvp)
                
                # Return all created/updated RSVPs
                return Response(
                    RSVPSerializer(created_rsvps, many=True).data,
                    status=status.HTTP_201_CREATED if not existing_rsvp else status.HTTP_200_OK
                )
            
            elif event.rsvp_mode == 'ONE_TAP_ALL':
                # ONE_TAP_ALL mode: Create RSVP for all allowed sub-events
                # Get allowed sub-events for this guest
                if guest:
                    allowed_sub_events = SubEvent.objects.filter(
                        guest_invites__guest=guest,
                        is_removed=False,
                        rsvp_enabled=True
                    )
                else:
                    # Public event - get all public-visible sub-events
                    allowed_sub_events = SubEvent.objects.filter(
                        event=event,
                        is_public_visible=True,
                        is_removed=False,
                        rsvp_enabled=True
                    )
                
                if not allowed_sub_events.exists():
                    return Response(
                        {'error': 'No sub-events available for RSVP'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Create/update RSVP for each allowed sub-event
                created_rsvps = []
                for sub_event in allowed_sub_events:
                    # Check if RSVP already exists for this sub-event
                    existing_sub_rsvp = RSVP.objects.filter(
                        event=event,
                        phone=phone,
                        sub_event=sub_event,
                        is_removed=False
                    ).first()
                    
                    if existing_sub_rsvp:
                        # Update existing RSVP
                        for key, value in rsvp_data.items():
                            setattr(existing_sub_rsvp, key, value)
                        if guest:
                            existing_sub_rsvp.guest = guest
                        existing_sub_rsvp.save()
                        created_rsvps.append(existing_sub_rsvp)
                    else:
                        # Create new RSVP
                        rsvp = RSVP.objects.create(
                            event=event,
                            sub_event=sub_event,
                            guest=guest,
                            **rsvp_data
                        )
                        created_rsvps.append(rsvp)
                
                # Return all created/updated RSVPs
                return Response(
                    RSVPSerializer(created_rsvps, many=True).data,
                    status=status.HTTP_201_CREATED if not existing_rsvp else status.HTTP_200_OK
                )
        
        # SIMPLE event: Keep existing behavior (sub_event = NULL)
        if existing_rsvp:
            # Check if RSVP itself is removed (shouldn't happen due to filter, but double-check)
            if existing_rsvp.is_removed:
                return Response(
                    {'error': 'This RSVP has been removed. You cannot update it.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Update existing RSVP
            for key, value in rsvp_data.items():
                setattr(existing_rsvp, key, value)
            if guest:
                existing_rsvp.guest = guest
            existing_rsvp.save()
            return Response(RSVPSerializer(existing_rsvp).data, status=status.HTTP_200_OK)
        else:
            # Create new RSVP (sub_event will be NULL for SIMPLE events)
            rsvp = RSVP.objects.create(event=event, guest=guest, **rsvp_data)
            return Response(RSVPSerializer(rsvp).data, status=status.HTTP_201_CREATED)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_image(request, event_id):
    """
    Upload an image file to S3 and return the public URL
    
    Accepts: multipart/form-data with 'image' field
    URL: /api/events/{event_id}/upload-image/
    Returns: { 'url': 'https://...' }
    """
    # Validate event_id is an integer
    try:
        event_id = int(event_id)
    except (ValueError, TypeError):
        return Response(
            {'error': 'Invalid event_id. Must be an integer.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get event and validate ownership
    try:
        event = Event.objects.get(id=event_id)
    except Event.DoesNotExist:
        return Response(
            {'error': 'Event not found.'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Verify user owns the event
    if event.host != request.user:
        return Response(
            {'error': 'You do not have permission to upload images to this event.'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    if 'image' not in request.FILES:
        return Response(
            {'error': 'No image file provided. Please include an image file in the request.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    image_file = request.FILES['image']
    
    # Validate file size (max 5MB)
    max_size = 5 * 1024 * 1024  # 5MB
    if image_file.size > max_size:
        return Response(
            {'error': f'Image file is too large. Maximum size is 5MB, but received {image_file.size / 1024 / 1024:.2f}MB.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate file type
    allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp', 'image/gif']
    if image_file.content_type not in allowed_types:
        return Response(
            {'error': f'Invalid file type. Allowed types: {", ".join(allowed_types)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Upload to S3 or local storage with event_id
        uploaded_url = upload_to_s3(image_file, event_id=event_id)
        
        # If in DEBUG mode and URL is relative, make it absolute
        if settings.DEBUG and uploaded_url.startswith('/'):
            # Construct full URL using request
            scheme = request.scheme
            host = request.get_host()
            full_url = f"{scheme}://{host}{uploaded_url}"
            uploaded_url = full_url
        
        return Response(
            {'url': uploaded_url},
            status=status.HTTP_200_OK
        )
    except Exception as e:
        # Log the actual error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Image upload failed for event {event_id}: {str(e)}', exc_info=True)
        
        return Response(
            {'error': f'Failed to upload image: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class SubEventViewSet(viewsets.ModelViewSet):
    """
    CRUD operations for sub-events - host only, privacy protected
    """
    serializer_class = SubEventSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_queryset(self):
        """Hosts can only see sub-events for their own events"""
        queryset = SubEvent.objects.filter(event__host=self.request.user, is_removed=False)
        
        # Filter by event_id if provided in URL kwargs (for /envelopes/<event_id>/sub-events/ endpoint)
        event_id = self.kwargs.get('event_id')
        if event_id:
            # Verify the event belongs to the user
            try:
                event = Event.objects.get(id=event_id, host=self.request.user)
                queryset = queryset.filter(event=event)
            except Event.DoesNotExist:
                from rest_framework.exceptions import NotFound
                raise NotFound("Event not found or you don't have permission to access it.")
        
        return queryset.order_by('start_at')
    
    def get_serializer_class(self):
        if self.action == 'create':
            return SubEventCreateSerializer
        return SubEventSerializer
    
    def get_object(self):
        """Override to ensure host can only access their own sub-events"""
        obj = super().get_object()
        if obj.event.host != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access sub-events for your own events.")
        return obj
    
    def perform_create(self, serializer):
        """Ensure sub-event is created for an event owned by the user"""
        event_id = self.kwargs.get('event_id')
        if not event_id:
            from rest_framework.exceptions import ValidationError
            raise ValidationError("event_id is required")
        
        event = get_object_or_404(Event, id=event_id, host=self.request.user)
        
        # Upgrade event to ENVELOPE if needed
        event.upgrade_to_envelope_if_needed()
        
        sub_event = serializer.save(event=event)
        
        # If rsvp_mode is PER_SUBEVENT, ensure event is ENVELOPE
        if event.rsvp_mode == 'PER_SUBEVENT':
            event.event_structure = 'ENVELOPE'
            event.save(update_fields=['event_structure', 'updated_at'])
        
        return sub_event
    
    def create(self, request, *args, **kwargs):
        """Create sub-event for an event"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        sub_event = self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(SubEventSerializer(sub_event).data, status=status.HTTP_201_CREATED, headers=headers)
    
    def perform_destroy(self, instance):
        """Soft delete sub-event"""
        instance.is_removed = True
        instance.save(update_fields=['is_removed', 'updated_at'])


class GuestInviteViewSet(viewsets.ModelViewSet):
    """
    Manage guest sub-event assignments - host only, privacy protected
    """
    serializer_class = GuestSubEventInviteSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_queryset(self):
        """Hosts can only see guest invites for their own events"""
        return GuestSubEventInvite.objects.filter(
            guest__event__host=self.request.user
        )
    
    def get_object(self):
        """Override to ensure host can only access their own guest invites"""
        obj = super().get_object()
        if obj.guest.event.host != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access guest invites for your own events.")
        return obj
    
    @action(detail=False, methods=['get'], url_path='event/(?P<event_id>[^/.]+)')
    def by_event(self, request, event_id=None):
        """Get all guest invites for an event"""
        event = get_object_or_404(Event, id=event_id, host=request.user)
        
        # Get all guests with their sub-event assignments
        guests = Guest.objects.filter(event=event, is_removed=False)
        result = []
        
        for guest in guests:
            invites = GuestSubEventInvite.objects.filter(guest=guest)
            result.append({
                'guest': GuestSerializer(guest).data,
                'sub_event_ids': list(invites.values_list('sub_event_id', flat=True))
            })
        
        return Response(result)
    
    @action(detail=False, methods=['put'], url_path='guest/(?P<guest_id>[^/.]+)')
    def update_guest_invites(self, request, guest_id=None):
        """Update sub-event assignments for a guest"""
        guest = get_object_or_404(Guest, id=guest_id, event__host=request.user)
        
        # Verify ownership
        if guest.event.host != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only update guest invites for your own events.")
        
        # Get sub_event_ids from request
        sub_event_ids = request.data.get('sub_event_ids', [])
        if not isinstance(sub_event_ids, list):
            return Response(
                {'error': 'sub_event_ids must be a list'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify all sub-events belong to the same event
        sub_events = SubEvent.objects.filter(
            id__in=sub_event_ids,
            event=guest.event,
            is_removed=False
        )
        
        if sub_events.count() != len(sub_event_ids):
            return Response(
                {'error': 'Some sub-events not found or do not belong to this event'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Upgrade event to ENVELOPE if needed
        guest.event.upgrade_to_envelope_if_needed()
        
        # Remove existing invites
        GuestSubEventInvite.objects.filter(guest=guest).delete()
        
        # Create new invites
        for sub_event in sub_events:
            GuestSubEventInvite.objects.get_or_create(
                guest=guest,
                sub_event=sub_event
            )
        
        # Generate guest token if not present
        guest.generate_guest_token()
        
        # Return updated guest data
        return Response({
            'guest': GuestSerializer(guest).data,
            'sub_event_ids': list(GuestSubEventInvite.objects.filter(guest=guest).values_list('sub_event_id', flat=True))
        })


class WhatsAppTemplateViewSet(viewsets.ModelViewSet):
    """ViewSet for managing WhatsApp templates per event"""
    serializer_class = WhatsAppTemplateSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_queryset(self):
        """Filter templates by event and verify event ownership"""
        # Start with all templates owned by the user (allows detail operations without event_id)
        queryset = WhatsAppTemplate.objects.filter(event__host=self.request.user)
        
        # Filter by event_id if provided in URL kwargs (for nested list/create routes)
        event_id = self.kwargs.get('event_id') or self.request.query_params.get('event_id') or self.request.data.get('event_id')
        
        if event_id:
            # Verify the event belongs to the user and filter
            try:
                event = Event.objects.get(id=event_id, host=self.request.user)
                queryset = queryset.filter(event=event)
            except Event.DoesNotExist:
                return WhatsAppTemplate.objects.none()
        
        return queryset
    
    def get_object(self):
        """Override to verify event ownership"""
        obj = super().get_object()
        # Verify the event belongs to the user
        if obj.event.host != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access templates for your own events.")
        return obj
    
    def perform_create(self, serializer):
        """Ensure template is created with valid event owned by user"""
        # Get event_id from URL kwargs (for nested routes) or request data
        event_id = self.kwargs.get('event_id') or self.request.data.get('event')
        if not event_id:
            raise drf_serializers.ValidationError({'event': 'Event ID is required.'})
        
        try:
            event = Event.objects.get(id=event_id, host=self.request.user)
        except Event.DoesNotExist:
            raise drf_serializers.ValidationError({'event': 'Event not found or you do not have permission.'})
        
        # Check for duplicate name
        if WhatsAppTemplate.objects.filter(event=event, name=serializer.validated_data['name']).exists():
            raise drf_serializers.ValidationError({'name': 'A template with this name already exists for this event.'})
        
        # Set created_by
        serializer.save(event=event, created_by=self.request.user)
        
        # Handle is_default flag
        if serializer.validated_data.get('is_default', False):
            # Unset other defaults for this event
            WhatsAppTemplate.objects.filter(event=event, is_default=True).exclude(id=serializer.instance.id).update(is_default=False)
    
    def perform_update(self, serializer):
        """Ensure template update maintains event ownership"""
        event_id = serializer.validated_data.get('event') or self.get_object().event.id
        
        try:
            event = Event.objects.get(id=event_id, host=self.request.user)
        except Event.DoesNotExist:
            raise drf_serializers.ValidationError({'event': 'Event not found or you do not have permission.'})
        
        # Check for duplicate name (excluding current template)
        template = self.get_object()
        if WhatsAppTemplate.objects.filter(event=event, name=serializer.validated_data['name']).exclude(id=template.id).exists():
            raise drf_serializers.ValidationError({'name': 'A template with this name already exists for this event.'})
        
        serializer.save()
        
        # Handle is_default flag
        if serializer.validated_data.get('is_default', False):
            # Unset other defaults for this event
            WhatsAppTemplate.objects.filter(event=event, is_default=True).exclude(id=template.id).update(is_default=False)
    
    @action(detail=True, methods=['post'])
    def preview(self, request, id=None):
        """Generate preview with sample data"""
        template = self.get_object()
        sample_data = request.data.get('sample_data', {})
        preview_text = template.get_preview(sample_data)
        return Response({
            'preview': preview_text,
            'template': WhatsAppTemplateSerializer(template).data
        })
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, id=None):
        """Create a copy of this template"""
        template = self.get_object()
        new_name = request.data.get('name') or f"{template.name} (Copy)"
        
        # Check if name already exists
        if WhatsAppTemplate.objects.filter(event=template.event, name=new_name).exists():
            return Response(
                {'error': f'A template with the name "{new_name}" already exists for this event.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        new_template = WhatsAppTemplate.objects.create(
            event=template.event,
            name=new_name,
            message_type=template.message_type,
            template_text=template.template_text,
            description=template.description,
            is_active=True,
            usage_count=0,
            last_used_at=None
        )
        
        return Response(
            WhatsAppTemplateSerializer(new_template).data,
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=True, methods=['post'])
    def archive(self, request, id=None):
        """Archive template (set is_active=False)"""
        template = self.get_object()
        template.is_active = False
        template.save(update_fields=['is_active'])
        return Response(WhatsAppTemplateSerializer(template).data)
    
    @action(detail=True, methods=['post'])
    def activate(self, request, id=None):
        """Activate template (set is_active=True)"""
        template = self.get_object()
        template.is_active = True
        template.save(update_fields=['is_active'])
        return Response(WhatsAppTemplateSerializer(template).data)
    
    @action(detail=True, methods=['post'])
    def increment_usage(self, request, id=None):
        """Increment usage count and update last_used_at"""
        template = self.get_object()
        template.increment_usage()
        return Response(WhatsAppTemplateSerializer(template).data)
    
    def perform_destroy(self, instance):
        """Prevent deletion of system default templates"""
        if instance.is_system_default:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("System default templates cannot be deleted.")
        super().perform_destroy(instance)
    
    @action(detail=True, methods=['post'], url_path='set-default')
    def set_default(self, request, id=None):
        """Set this template as the event's default template"""
        template = self.get_object()
        
        # Unset other defaults for this event
        WhatsAppTemplate.objects.filter(event=template.event, is_default=True).exclude(id=template.id).update(is_default=False)
        
        # Set this template as default
        template.is_default = True
        template.save(update_fields=['is_default'])
        
        return Response(WhatsAppTemplateSerializer(template).data)
    
    @action(detail=True, methods=['post'], url_path='preview-with-guest')
    def preview_with_guest(self, request, id=None):
        """Preview template with specific guest data"""
        template = self.get_object()
        guest_id = request.data.get('guest_id')
        
        if not guest_id:
            return Response({'error': 'guest_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            guest = Guest.objects.get(id=guest_id, event=template.event)
        except Guest.DoesNotExist:
            return Response({'error': 'Guest not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get base URL from request
        base_url = request.build_absolute_uri('/').rstrip('/')
        
        # Render template with guest data
        from .utils import render_template_with_guest
        rendered_message, warnings = render_template_with_guest(
            template.template_text,
            template.event,
            guest,
            base_url
        )
        
        return Response({
            'preview': rendered_message,
            'warnings': warnings,
            'template': WhatsAppTemplateSerializer(template).data,
            'guest': GuestSerializer(guest).data,
        })


# Standalone view functions for WhatsApp template actions
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def whatsapp_template_preview(request, id):
    """Preview template with sample data"""
    try:
        template = WhatsAppTemplate.objects.get(id=id)
        # Verify ownership
        if template.event.host != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access templates for your own events.")
        
        sample_data = request.data.get('sample_data', {})
        preview_text = template.get_preview(sample_data)
        return Response({
            'preview': preview_text,
            'template': WhatsAppTemplateSerializer(template).data
        })
    except WhatsAppTemplate.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def whatsapp_template_duplicate(request, id):
    """Duplicate a template"""
    try:
        template = WhatsAppTemplate.objects.get(id=id)
        # Verify ownership
        if template.event.host != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access templates for your own events.")
        
        new_name = request.data.get('name') or f"{template.name} (Copy)"
        
        # Check if name already exists
        if WhatsAppTemplate.objects.filter(event=template.event, name=new_name).exists():
            return Response(
                {'error': f'A template with the name "{new_name}" already exists for this event.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        new_template = WhatsAppTemplate.objects.create(
            event=template.event,
            name=new_name,
            message_type=template.message_type,
            template_text=template.template_text,
            description=template.description,
            is_active=True,
            usage_count=0,
            last_used_at=None
        )
        
        return Response(
            WhatsAppTemplateSerializer(new_template).data,
            status=status.HTTP_201_CREATED
        )
    except WhatsAppTemplate.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def whatsapp_template_archive(request, id):
    """Archive a template"""
    try:
        template = WhatsAppTemplate.objects.get(id=id)
        # Verify ownership
        if template.event.host != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access templates for your own events.")
        
        template.is_active = False
        template.save(update_fields=['is_active'])
        return Response(WhatsAppTemplateSerializer(template).data)
    except WhatsAppTemplate.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def whatsapp_template_activate(request, id):
    """Activate a template"""
    try:
        template = WhatsAppTemplate.objects.get(id=id)
        # Verify ownership
        if template.event.host != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access templates for your own events.")
        
        template.is_active = True
        template.save(update_fields=['is_active'])
        return Response(WhatsAppTemplateSerializer(template).data)
    except WhatsAppTemplate.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def whatsapp_template_increment_usage(request, id):
    """Increment template usage"""
    try:
        template = WhatsAppTemplate.objects.get(id=id)
        # Verify ownership
        if template.event.host != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You can only access templates for your own events.")
        
        template.increment_usage()
        return Response(WhatsAppTemplateSerializer(template).data)
    except WhatsAppTemplate.DoesNotExist:
        return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)


